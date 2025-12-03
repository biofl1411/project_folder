#!/usr/bin/env python
# -*- coding: utf-8 -*-

'''
사용자 관리 탭 - 관리자 전용
'''

from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                           QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
                           QHeaderView, QMessageBox, QComboBox, QGroupBox,
                           QFormLayout, QCheckBox, QScrollArea, QFrame, QGridLayout,
                           QFileDialog)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QFont

from models.users import (User, DEPARTMENTS, PERMISSION_LABELS, DEFAULT_PASSWORD,
                         PERMISSION_CATEGORIES, PERMISSION_BY_CATEGORY,
                         PERMISSION_DESCRIPTIONS, get_default_permissions)

import os


class UserManagementTab(QWidget):
    def __init__(self, current_user=None):
        super().__init__()
        self.current_user = current_user
        self.selected_user_id = None
        self.permission_checkboxes = {}
        self.category_checkboxes = {}  # 카테고리 전체선택 체크박스
        self.initUI()
        self.load_users()

    def set_current_user(self, user):
        """현재 로그인한 사용자 설정"""
        self.current_user = user

    def initUI(self):
        """UI 초기화"""
        layout = QHBoxLayout(self)
        layout.setSpacing(15)

        # 왼쪽: 사용자 목록
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)

        # 사용자 목록 라벨
        list_label = QLabel("사용자 목록")
        list_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        left_layout.addWidget(list_label)

        # 사용자 테이블
        self.user_table = QTableWidget()
        self.user_table.setColumnCount(7)
        self.user_table.setHorizontalHeaderLabels(['ID', '아이디', '이름', '부서', '이메일', '마지막 로그인', '활성화'])
        self.user_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.user_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.user_table.setSelectionMode(QTableWidget.SingleSelection)
        self.user_table.itemSelectionChanged.connect(self.on_user_selected)
        self.user_table.cellDoubleClicked.connect(self.on_cell_double_clicked)
        self.user_table.setColumnHidden(0, True)  # ID 컬럼 숨김
        left_layout.addWidget(self.user_table)

        # 버튼 영역 1 (기본 기능)
        btn_layout = QHBoxLayout()

        refresh_btn = QPushButton("새로고침")
        refresh_btn.clicked.connect(self.load_users)
        btn_layout.addWidget(refresh_btn)

        delete_btn = QPushButton("삭제")
        delete_btn.setStyleSheet("background-color: #f44336; color: white;")
        delete_btn.clicked.connect(self.delete_user)
        btn_layout.addWidget(delete_btn)

        reset_pwd_btn = QPushButton("비밀번호 초기화")
        reset_pwd_btn.clicked.connect(self.reset_password)
        btn_layout.addWidget(reset_pwd_btn)

        left_layout.addLayout(btn_layout)

        # 버튼 영역 2 (엑셀 기능)
        excel_btn_layout = QHBoxLayout()

        import_btn = QPushButton("엑셀 불러오기")
        import_btn.setStyleSheet("background-color: #2196F3; color: white;")
        import_btn.clicked.connect(self.import_excel)
        excel_btn_layout.addWidget(import_btn)

        template_btn = QPushButton("엑셀 양식 다운로드")
        template_btn.setStyleSheet("background-color: #4CAF50; color: white;")
        template_btn.clicked.connect(self.download_template)
        excel_btn_layout.addWidget(template_btn)

        left_layout.addLayout(excel_btn_layout)

        layout.addWidget(left_widget, 1)

        # 오른쪽: 사용자 추가/수정
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)

        # 사용자 정보 입력 그룹
        info_group = QGroupBox("사용자 정보")
        info_layout = QFormLayout(info_group)

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("사용자 이름")
        info_layout.addRow("이름:", self.name_input)

        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("로그인 아이디")
        info_layout.addRow("아이디:", self.username_input)

        self.department_combo = QComboBox()
        self.department_combo.addItems([d for d in DEPARTMENTS if d != '관리자'])
        info_layout.addRow("부서:", self.department_combo)

        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("이메일 주소 (메일 발송에 사용)")
        info_layout.addRow("이메일:", self.email_input)

        # 초기 비밀번호 안내
        pwd_label = QLabel(f"초기 비밀번호: {DEFAULT_PASSWORD}")
        pwd_label.setStyleSheet("color: #666; font-size: 10px;")
        info_layout.addRow("", pwd_label)

        right_layout.addWidget(info_group)

        # 권한 설정 그룹
        perm_group = QGroupBox("권한 설정")
        perm_layout = QVBoxLayout(perm_group)

        # 전체 선택/해제 버튼
        all_btn_layout = QHBoxLayout()
        select_all_btn = QPushButton("전체 선택")
        select_all_btn.clicked.connect(lambda: self.toggle_all_permissions(True))
        all_btn_layout.addWidget(select_all_btn)

        deselect_all_btn = QPushButton("전체 해제")
        deselect_all_btn.clicked.connect(lambda: self.toggle_all_permissions(False))
        all_btn_layout.addWidget(deselect_all_btn)
        perm_layout.addLayout(all_btn_layout)

        # 스크롤 영역
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        perm_widget = QWidget()
        perm_main_layout = QVBoxLayout(perm_widget)
        perm_main_layout.setSpacing(10)

        # 카테고리별 권한 그룹 생성
        for cat_key, cat_name in PERMISSION_CATEGORIES.items():
            cat_group = QGroupBox(cat_name)
            cat_group.setStyleSheet("""
                QGroupBox {
                    font-weight: bold;
                    border: 1px solid #ccc;
                    border-radius: 5px;
                    margin-top: 10px;
                    padding-top: 10px;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 5px;
                }
            """)
            cat_layout = QVBoxLayout(cat_group)

            # 카테고리 전체 선택 체크박스
            cat_header = QHBoxLayout()
            cat_all_checkbox = QCheckBox("전체")
            cat_all_checkbox.setStyleSheet("font-weight: bold; color: #2196F3;")
            cat_all_checkbox.stateChanged.connect(
                lambda state, ck=cat_key: self.toggle_category_permissions(ck, state == Qt.Checked)
            )
            self.category_checkboxes[cat_key] = cat_all_checkbox
            cat_header.addWidget(cat_all_checkbox)
            cat_header.addStretch()
            cat_layout.addLayout(cat_header)

            # 권한 체크박스들
            perm_keys = PERMISSION_BY_CATEGORY.get(cat_key, [])
            perm_grid = QGridLayout()
            perm_grid.setSpacing(5)

            row = 0
            col = 0
            for perm_key in perm_keys:
                perm_label_text = PERMISSION_LABELS.get(perm_key, perm_key)
                perm_desc = PERMISSION_DESCRIPTIONS.get(perm_key, '')

                # 체크박스 + 도움말 아이콘을 담을 위젯
                item_widget = QWidget()
                item_layout = QHBoxLayout(item_widget)
                item_layout.setContentsMargins(0, 0, 0, 0)
                item_layout.setSpacing(3)

                checkbox = QCheckBox(perm_label_text)
                checkbox.setStyleSheet("""
                    QCheckBox::indicator {
                        width: 16px;
                        height: 16px;
                    }
                    QCheckBox::indicator:unchecked {
                        background-color: #ffcccc;
                        border: 1px solid #999;
                        border-radius: 3px;
                    }
                    QCheckBox::indicator:checked {
                        background-color: #ccffcc;
                        border: 1px solid #999;
                        border-radius: 3px;
                    }
                """)
                checkbox.stateChanged.connect(
                    lambda state, ck=cat_key: self.update_category_checkbox(ck)
                )

                # 도움말 아이콘 (?)
                help_label = QLabel("?")
                help_label.setStyleSheet("""
                    QLabel {
                        color: white;
                        background-color: #3498db;
                        border-radius: 8px;
                        font-size: 10px;
                        font-weight: bold;
                        padding: 1px 5px;
                    }
                    QLabel:hover {
                        background-color: #2980b9;
                    }
                """)
                help_label.setToolTip(perm_desc)
                help_label.setFixedSize(16, 16)
                help_label.setAlignment(Qt.AlignCenter)

                item_layout.addWidget(checkbox)
                item_layout.addWidget(help_label)
                item_layout.addStretch()

                self.permission_checkboxes[perm_key] = checkbox
                perm_grid.addWidget(item_widget, row, col)

                col += 1
                if col >= 2:  # 2열로 배치
                    col = 0
                    row += 1

            cat_layout.addLayout(perm_grid)
            perm_main_layout.addWidget(cat_group)

        perm_main_layout.addStretch()
        scroll.setWidget(perm_widget)
        perm_layout.addWidget(scroll)

        right_layout.addWidget(perm_group)

        # 저장 버튼
        save_btn_layout = QHBoxLayout()

        clear_btn = QPushButton("새 사용자")
        clear_btn.clicked.connect(self.clear_form)
        save_btn_layout.addWidget(clear_btn)

        save_btn = QPushButton("저장")
        save_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        save_btn.clicked.connect(self.save_user)
        save_btn_layout.addWidget(save_btn)

        right_layout.addLayout(save_btn_layout)

        layout.addWidget(right_widget, 1)

    def toggle_all_permissions(self, checked):
        """모든 권한 전체 선택/해제"""
        for checkbox in self.permission_checkboxes.values():
            checkbox.setChecked(checked)

    def toggle_category_permissions(self, category_key, checked):
        """카테고리별 권한 전체 선택/해제"""
        perm_keys = PERMISSION_BY_CATEGORY.get(category_key, [])
        for perm_key in perm_keys:
            if perm_key in self.permission_checkboxes:
                self.permission_checkboxes[perm_key].setChecked(checked)

    def update_category_checkbox(self, category_key):
        """카테고리 내 개별 체크박스 상태에 따라 전체 체크박스 업데이트"""
        perm_keys = PERMISSION_BY_CATEGORY.get(category_key, [])
        all_checked = all(
            self.permission_checkboxes.get(pk, QCheckBox()).isChecked()
            for pk in perm_keys
        )

        if category_key in self.category_checkboxes:
            # 시그널 순환 방지
            self.category_checkboxes[category_key].blockSignals(True)
            self.category_checkboxes[category_key].setChecked(all_checked)
            self.category_checkboxes[category_key].blockSignals(False)

    def load_users(self):
        """사용자 목록 로드"""
        self.user_table.setRowCount(0)
        users = User.get_all()

        for user in users:
            # admin 계정은 목록에서 제외 (수정/삭제 방지)
            if user.get('username') == 'admin':
                continue

            row = self.user_table.rowCount()
            self.user_table.insertRow(row)

            self.user_table.setItem(row, 0, QTableWidgetItem(str(user.get('id', ''))))
            self.user_table.setItem(row, 1, QTableWidgetItem(user.get('username', '')))
            self.user_table.setItem(row, 2, QTableWidgetItem(user.get('name', '')))
            self.user_table.setItem(row, 3, QTableWidgetItem(user.get('department', '')))
            self.user_table.setItem(row, 4, QTableWidgetItem(user.get('email', '') or ''))
            self.user_table.setItem(row, 5, QTableWidgetItem(user.get('last_login', '') or ''))

            # 활성화 상태 (O/X)
            is_active = user.get('is_active', 0)
            active_item = QTableWidgetItem('O' if is_active else 'X')
            active_item.setTextAlignment(Qt.AlignCenter)
            if is_active:
                active_item.setForeground(QColor('#4CAF50'))  # 녹색
            else:
                active_item.setForeground(QColor('#f44336'))  # 빨간색
            active_item.setFont(QFont('Arial', 12, QFont.Bold))
            self.user_table.setItem(row, 6, active_item)

    def on_user_selected(self):
        """사용자 선택 시"""
        selected = self.user_table.selectedItems()
        if not selected:
            return

        row = selected[0].row()
        user_id = int(self.user_table.item(row, 0).text())
        self.selected_user_id = user_id

        user = User.get_by_id(user_id)
        if user:
            self.name_input.setText(user.get('name', ''))
            self.username_input.setText(user.get('username', ''))
            self.username_input.setEnabled(False)  # 아이디는 수정 불가

            department = user.get('department', '')
            index = self.department_combo.findText(department)
            if index >= 0:
                self.department_combo.setCurrentIndex(index)

            self.email_input.setText(user.get('email', '') or '')

            # 권한 체크박스 설정
            permissions = user.get('permissions', {})
            for perm_key, checkbox in self.permission_checkboxes.items():
                checkbox.setChecked(permissions.get(perm_key, False))

            # 카테고리 체크박스 업데이트
            for cat_key in PERMISSION_CATEGORIES.keys():
                self.update_category_checkbox(cat_key)

    def clear_form(self):
        """입력 폼 초기화"""
        self.selected_user_id = None
        self.name_input.clear()
        self.username_input.clear()
        self.username_input.setEnabled(True)
        self.department_combo.setCurrentIndex(0)
        self.email_input.clear()

        # 권한 체크박스 초기화 (모두 해제)
        for checkbox in self.permission_checkboxes.values():
            checkbox.setChecked(False)

        # 카테고리 체크박스 업데이트
        for cat_key in PERMISSION_CATEGORIES.keys():
            self.update_category_checkbox(cat_key)

        self.user_table.clearSelection()

    def save_user(self):
        """사용자 저장"""
        name = self.name_input.text().strip()
        username = self.username_input.text().strip()
        department = self.department_combo.currentText()
        email = self.email_input.text().strip()

        if not name:
            QMessageBox.warning(self, "입력 오류", "이름을 입력하세요.")
            return

        if not username:
            QMessageBox.warning(self, "입력 오류", "아이디를 입력하세요.")
            return

        # 권한 수집
        permissions = {}
        for perm_key, checkbox in self.permission_checkboxes.items():
            permissions[perm_key] = checkbox.isChecked()

        if self.selected_user_id:
            # 기존 사용자 수정
            success = User.update(
                self.selected_user_id,
                name=name,
                department=department,
                permissions=permissions,
                email=email
            )
            if success:
                QMessageBox.information(self, "성공", "사용자 정보가 수정되었습니다.")
                self.load_users()
            else:
                QMessageBox.critical(self, "오류", "사용자 수정에 실패했습니다.")
        else:
            # 새 사용자 추가
            user_id = User.create(
                username=username,
                password=DEFAULT_PASSWORD,
                name=name,
                role='user',
                department=department,
                permissions=permissions,
                email=email
            )
            if user_id:
                QMessageBox.information(self, "성공",
                    f"사용자가 추가되었습니다.\n초기 비밀번호: {DEFAULT_PASSWORD}")
                self.clear_form()
                self.load_users()
            else:
                QMessageBox.critical(self, "오류", "사용자 추가에 실패했습니다.\n(아이디 중복 확인)")

    def delete_user(self):
        """사용자 삭제"""
        if not self.selected_user_id:
            QMessageBox.warning(self, "선택 오류", "삭제할 사용자를 선택하세요.")
            return

        reply = QMessageBox.question(
            self, "삭제 확인",
            "선택한 사용자를 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            if User.delete(self.selected_user_id):
                QMessageBox.information(self, "성공", "사용자가 삭제되었습니다.")
                self.clear_form()
                self.load_users()
            else:
                QMessageBox.critical(self, "오류", "사용자 삭제에 실패했습니다.")

    def reset_password(self):
        """비밀번호 초기화"""
        if not self.selected_user_id:
            QMessageBox.warning(self, "선택 오류", "비밀번호를 초기화할 사용자를 선택하세요.")
            return

        reply = QMessageBox.question(
            self, "비밀번호 초기화",
            f"선택한 사용자의 비밀번호를 초기화하시겠습니까?\n새 비밀번호: {DEFAULT_PASSWORD}",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            if User.reset_password(self.selected_user_id):
                QMessageBox.information(self, "성공",
                    f"비밀번호가 초기화되었습니다.\n새 비밀번호: {DEFAULT_PASSWORD}")
                self.load_users()  # 목록 새로고침
            else:
                QMessageBox.critical(self, "오류", "비밀번호 초기화에 실패했습니다.")

    def on_cell_double_clicked(self, row, column):
        """셀 더블클릭 시 - 활성화 컬럼(6)만 처리"""
        if column != 6:  # 활성화 컬럼이 아니면 무시
            return

        # 사용자 ID 가져오기
        user_id_item = self.user_table.item(row, 0)
        if not user_id_item:
            return

        user_id = int(user_id_item.text())
        current_status = self.user_table.item(row, 6).text()
        is_currently_active = current_status == 'O'

        if is_currently_active:
            # 비활성화
            reply = QMessageBox.question(
                self, "사용자 비활성화",
                "선택한 사용자를 비활성화하시겠습니까?\n비활성화된 사용자는 로그인할 수 없습니다.",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                if User.toggle_active(user_id, activate=False):
                    QMessageBox.information(self, "성공", "사용자가 비활성화되었습니다.")
                    self.load_users()  # 즉시 새로고침
                else:
                    QMessageBox.critical(self, "오류", "비활성화에 실패했습니다.")
        else:
            # 활성화
            reply = QMessageBox.question(
                self, "사용자 활성화",
                f"선택한 사용자를 활성화하시겠습니까?\n비밀번호가 초기값({DEFAULT_PASSWORD})으로 설정됩니다.",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                if User.toggle_active(user_id, activate=True):
                    QMessageBox.information(self, "성공",
                        f"사용자가 활성화되었습니다.\n초기 비밀번호: {DEFAULT_PASSWORD}")
                    self.load_users()  # 즉시 새로고침
                else:
                    QMessageBox.critical(self, "오류", "활성화에 실패했습니다.")

    def download_template(self):
        """엑셀 양식 다운로드"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # 저장 경로 선택
            file_path, _ = QFileDialog.getSaveFileName(
                self, "엑셀 양식 저장",
                "사용자_등록_양식.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "사용자 등록"

            # 헤더 스타일
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 헤더 작성
            headers = ['아이디*', '이름*', '부서', '메일주소', '직통번호']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # 예시 데이터
            examples = [
                ['user001', '홍길동', '이화학팀', 'hong@example.com', '02-1234-5678'],
                ['user002', '김철수', '미생물팀', 'kim@example.com', '02-2345-6789'],
                ['user003', '이영희', '고객관리팀', 'lee@example.com', '02-3456-7890'],
            ]

            for row_idx, example in enumerate(examples, 2):
                for col_idx, value in enumerate(example, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border

            # 열 너비 조정
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 15

            # 부서 목록 시트 추가
            ws2 = wb.create_sheet(title="부서목록")
            ws2.cell(row=1, column=1, value="사용 가능한 부서 목록")
            ws2['A1'].font = Font(bold=True)
            for idx, dept in enumerate([d for d in DEPARTMENTS if d != '관리자'], 2):
                ws2.cell(row=idx, column=1, value=dept)

            # 안내 시트 추가
            ws3 = wb.create_sheet(title="안내사항")
            instructions = [
                "[ 사용자 일괄 등록 양식 안내 ]",
                "",
                "1. '사용자 등록' 시트에 데이터를 입력하세요.",
                "2. 아이디와 이름은 필수 입력 항목입니다. (*표시)",
                "3. 부서는 '부서목록' 시트를 참고하세요.",
                "4. 아이디는 영문/숫자 조합을 권장합니다.",
                f"5. 초기 비밀번호는 '{DEFAULT_PASSWORD}'로 설정됩니다.",
                "6. 메일주소와 직통번호는 선택 항목입니다.",
                "7. 예시 데이터(2~4행)는 삭제 후 실제 데이터를 입력하세요.",
            ]
            for idx, text in enumerate(instructions, 1):
                ws3.cell(row=idx, column=1, value=text)
            ws3.column_dimensions['A'].width = 60

            wb.save(file_path)
            QMessageBox.information(self, "성공", f"엑셀 양식이 저장되었습니다.\n{file_path}")

        except ImportError:
            QMessageBox.critical(self, "오류", "openpyxl 라이브러리가 설치되지 않았습니다.\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"양식 다운로드 중 오류 발생: {str(e)}")

    def import_excel(self):
        """엑셀 파일에서 사용자 일괄 등록"""
        try:
            import openpyxl

            # 파일 선택
            file_path, _ = QFileDialog.getOpenFileName(
                self, "엑셀 파일 선택",
                "",
                "Excel Files (*.xlsx *.xls)"
            )

            if not file_path:
                return

            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # 헤더 확인 (첫 번째 행)
            headers = [cell.value for cell in ws[1]]

            # 필수 컬럼 확인
            required = ['아이디', '이름']
            id_col = None
            name_col = None
            dept_col = None
            email_col = None
            phone_col = None

            for idx, header in enumerate(headers):
                if header and '아이디' in str(header):
                    id_col = idx
                elif header and '이름' in str(header):
                    name_col = idx
                elif header and '부서' in str(header):
                    dept_col = idx
                elif header and ('메일' in str(header) or 'mail' in str(header).lower()):
                    email_col = idx
                elif header and ('직통' in str(header) or '번호' in str(header) or 'phone' in str(header).lower()):
                    phone_col = idx

            if id_col is None or name_col is None:
                QMessageBox.warning(self, "형식 오류",
                    "엑셀 파일에 '아이디'와 '이름' 컬럼이 필요합니다.")
                return

            # 데이터 읽기 및 등록
            success_count = 0
            fail_count = 0
            fail_list = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                username = str(row[id_col]).strip() if row[id_col] else ''
                name = str(row[name_col]).strip() if row[name_col] else ''
                department = str(row[dept_col]).strip() if dept_col is not None and len(row) > dept_col and row[dept_col] else ''
                email = str(row[email_col]).strip() if email_col is not None and len(row) > email_col and row[email_col] else ''
                phone = str(row[phone_col]).strip() if phone_col is not None and len(row) > phone_col and row[phone_col] else ''

                # 빈 행 스킵
                if not username or not name:
                    continue

                # 부서 유효성 확인
                valid_depts = [d for d in DEPARTMENTS if d != '관리자']
                if department and department not in valid_depts:
                    department = valid_depts[0] if valid_depts else ''

                # 사용자 생성
                user_id = User.create(
                    username=username,
                    password=DEFAULT_PASSWORD,
                    name=name,
                    role='user',
                    department=department,
                    permissions=get_default_permissions(all_true=False),
                    email=email,
                    phone=phone
                )

                if user_id:
                    success_count += 1
                else:
                    fail_count += 1
                    fail_list.append(f"행 {row_idx}: {username} ({name})")

            # 결과 메시지
            msg = f"등록 완료: {success_count}명"
            if fail_count > 0:
                msg += f"\n등록 실패: {fail_count}명 (아이디 중복 등)"
                if fail_list[:5]:  # 최대 5개만 표시
                    msg += "\n\n실패 목록:\n" + "\n".join(fail_list[:5])
                    if len(fail_list) > 5:
                        msg += f"\n... 외 {len(fail_list) - 5}건"

            QMessageBox.information(self, "엑셀 불러오기 완료", msg)
            self.load_users()

        except ImportError:
            QMessageBox.critical(self, "오류", "openpyxl 라이브러리가 설치되지 않았습니다.\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 불러오기 중 오류 발생: {str(e)}")
